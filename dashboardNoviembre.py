import streamlit as st
st.set_page_config(layout="wide", page_icon="🏦", page_title="AFP Noviembre 2025")
# Forzar fondo blanco en toda la app
st.markdown("""
<style>
body, .main, [data-testid="stAppViewContainer"], [data-testid="stAppViewBlockContainer"] {
    background: #fff !important;
}
</style>
""", unsafe_allow_html=True)
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import re
import os
from datetime import datetime

# Ruta del archivo Excel (usar ruta relativa para Streamlit Cloud)
EXCEL_PATH = os.path.join(os.getcwd(), "DATA TOTAL WORLDTEL NOVIEMBRE 2025.xlsx")  # Confirmar ruta relativa

# Verificar si el archivo Excel existe
if not os.path.exists(EXCEL_PATH):
    st.error(f"El archivo Excel no se encuentra en la ruta especificada: {EXCEL_PATH}. Verifica que el archivo exista y que la ruta sea correcta.")
else:
    try:
        df = pd.read_excel(EXCEL_PATH)
    except Exception as e:
        st.error(f"Error al cargar el archivo Excel: {e}")

# Cargar datos

def load_data():
    df = pd.read_excel(EXCEL_PATH)
    # Verificar que la columna 'razon_social' esté presente
    if 'RAZON SOCIAL' not in df.columns:
        st.error("La columna 'RAZON SOCIAL' no está presente en el archivo Excel.")
    return df

df = load_data()

# Ocultar mensajes de verificación del archivo Excel y columnas disponibles
# st.write("Columnas disponibles en el DataFrame:", df.columns.tolist())
# st.success("Archivo Excel cargado correctamente.")

# Ajustar el nombre de la columna 'RAZON SOCIAL' al cargar los datos desde el archivo Excel
df['razon_social'] = df['RAZON SOCIAL'] if 'RAZON SOCIAL' in df.columns else ''

# Definir la función render_historial_pagos al inicio del archivo

def render_historial_pagos(df_pagos):
    """
    Renderiza el historial de pagos en la interfaz de Streamlit.
    Maneja las columnas FECHA DE PAGO P, REC. PLANILLAS, FECHA DE PAGO G, REC. GASTOS.
    """
    if df_pagos.empty:
        st.warning("No hay datos disponibles para mostrar en el historial de pagos.")
        return

    st.markdown("---")
    st.subheader("💰 Historial de Pagos Recibidos")

    # Filtros
    col_filtro1, col_filtro2, col_filtro3 = st.columns(3)

    campanas_pagos = ["Todas"] + sorted(df_pagos['campana'].dropna().unique().tolist())
    tipos_pago = ["Todos"] + sorted(df_pagos['tipo_pago'].dropna().unique().tolist())

    with col_filtro1:
        campana_pago_filter = st.selectbox("📋 Campaña (Pagos)", campanas_pagos, key="campana_pagos_historial")

    with col_filtro2:
        tipo_pago_filter = st.selectbox("💼 Tipo de Pago", tipos_pago, key="tipo_pagos_historial")

    # Ajustar el filtrado de fechas para incluir las más recientes
    rango_fechas = None
    if 'fecha' in df_pagos.columns and not df_pagos['fecha'].isna().all():
        fecha_min, fecha_max = df_pagos['fecha'].min().date(), df_pagos['fecha'].max().date()
        with col_filtro3:
            fecha_inicio = st.date_input("Fecha de Inicio", value=fecha_min, min_value=fecha_min, max_value=fecha_max, key="fecha_inicio_simple")
            fecha_fin = st.date_input("Fecha Fin", value=fecha_max, min_value=fecha_min, max_value=fecha_max, key="fecha_fin_simple")
            rango_fechas = (fecha_inicio, fecha_fin) if fecha_inicio <= fecha_fin else (fecha_fin, fecha_inicio)

    # Filtrar datos
    df_pagos_filtrado = df_pagos.copy()
    if campana_pago_filter != "Todas":
        df_pagos_filtrado = df_pagos_filtrado[df_pagos_filtrado['campana'] == campana_pago_filter]
    if tipo_pago_filter != "Todos":
        df_pagos_filtrado = df_pagos_filtrado[df_pagos_filtrado['tipo_pago'] == tipo_pago_filter]
    if rango_fechas:
        fecha_inicio, fecha_fin = rango_fechas
        df_pagos_filtrado = df_pagos_filtrado.loc[
            df_pagos_filtrado['fecha'].notna() &
            (df_pagos_filtrado['fecha'].dt.date >= fecha_inicio) &
            (df_pagos_filtrado['fecha'].dt.date <= fecha_fin)
        ]

    # Ordenar por fecha descendente para mostrar los pagos más recientes primero
    if 'fecha' in df_pagos_filtrado.columns:
        df_pagos_filtrado = df_pagos_filtrado.sort_values(by='fecha', ascending=False)

    # Métricas
    col_metric1, col_metric2, col_metric3, col_metric4, col_metric5 = st.columns(5)
    with col_metric1:
        st.metric("📊 Total Pagos", f"{len(df_pagos_filtrado):,}")
    with col_metric2:
        st.metric("🏦 Monto Planillas", f"S/. {df_pagos_filtrado[df_pagos_filtrado['tipo_pago'] == 'PLANILLAS']['monto'].sum():,.2f}")
    with col_metric3:
        st.metric("🏛️ Monto Gastos", f"S/. {df_pagos_filtrado[df_pagos_filtrado['tipo_pago'] == 'GASTOS']['monto'].sum():,.2f}")
    with col_metric4:
        st.metric("🏦 Pagos Planillas", f"{len(df_pagos_filtrado[df_pagos_filtrado['tipo_pago'] == 'PLANILLAS']):,}")
    with col_metric5:
        st.metric("🏛️ Pagos Gastos", f"{len(df_pagos_filtrado[df_pagos_filtrado['tipo_pago'] == 'GASTOS']):,}")

    # Gráficos
    if not df_pagos_filtrado.empty and 'fecha' in df_pagos_filtrado.columns:
        st.markdown("### 📈 Análisis de Pagos por Tipo")
        tab1, tab2, tab3 = st.tabs(["🏦 PAGOS PLANILLAS", "🏛️ PAGOS GASTOS", "📊 COMPARACIÓN"])

        with tab1:
            st.markdown("#### 🏦 Evolución de Pagos de Planillas por Día")
            df_planillas = df_pagos_filtrado[df_pagos_filtrado['tipo_pago'] == 'PLANILLAS']
            if not df_planillas.empty:
                df_planillas_clean = df_planillas[df_planillas['fecha'].notna()].copy()
                df_planillas_clean['fecha_dia'] = df_planillas_clean['fecha'].dt.date
                planillas_por_dia = df_planillas_clean.groupby('fecha_dia')['monto'].sum().reset_index()

                # Agregar selectbox para alternar entre evolución de pagos y recaudo acumulado
                tipo_grafico = st.selectbox("Tipo de gráfico", ["Evolución de Pagos", "Recaudo Acumulado"], key="tipo_grafico_planillas")

                if tipo_grafico == "Evolución de Pagos":
                    st.line_chart(planillas_por_dia.set_index('fecha_dia'))
                elif tipo_grafico == "Recaudo Acumulado":
                    planillas_por_dia['recaudo_acumulado'] = planillas_por_dia['monto'].cumsum()
                    st.line_chart(planillas_por_dia.set_index('fecha_dia')['recaudo_acumulado'])

        with tab2:
            st.markdown("#### 🏛️ Evolución de Pagos de Gastos por Día")
            df_gastos = df_pagos_filtrado[df_pagos_filtrado['tipo_pago'] == 'GASTOS']
            if not df_gastos.empty:
                df_gastos_clean = df_gastos[df_gastos['fecha'].notna()].copy()
                df_gastos_clean['fecha_dia'] = df_gastos_clean['fecha'].dt.date
                gastos_por_dia = df_gastos_clean.groupby('fecha_dia')['monto'].sum().reset_index()

                # Agregar selectbox para alternar entre evolución de pagos y recaudo acumulado
                tipo_grafico = st.selectbox("Tipo de gráfico", ["Evolución de Pagos", "Recaudo Acumulado"], key="tipo_grafico_gastos")

                if tipo_grafico == "Evolución de Pagos":
                    st.line_chart(gastos_por_dia.set_index('fecha_dia'))
                elif tipo_grafico == "Recaudo Acumulado":
                    gastos_por_dia['recaudo_acumulado'] = gastos_por_dia['monto'].cumsum()
                    st.line_chart(gastos_por_dia.set_index('fecha_dia')['recaudo_acumulado'])

        with tab3:
            st.markdown("#### 📊 Comparación: Planillas vs Gastos")
            df_planillas = df_pagos_filtrado[df_pagos_filtrado['tipo_pago'] == 'PLANILLAS']
            df_gastos = df_pagos_filtrado[df_pagos_filtrado['tipo_pago'] == 'GASTOS']

            if not df_planillas.empty or not df_gastos.empty:
                df_planillas_clean = df_planillas[df_planillas['fecha'].notna()].copy()
                df_planillas_clean['fecha_dia'] = df_planillas_clean['fecha'].dt.date
                planillas_por_dia = df_planillas_clean.groupby('fecha_dia')['monto'].sum().reset_index()
                planillas_por_dia['tipo'] = 'Planillas'

                df_gastos_clean = df_gastos[df_gastos['fecha'].notna()].copy()
                df_gastos_clean['fecha_dia'] = df_gastos_clean['fecha'].dt.date
                gastos_por_dia = df_gastos_clean.groupby('fecha_dia')['monto'].sum().reset_index()
                gastos_por_dia['tipo'] = 'Gastos'

                comparacion_df = pd.concat([planillas_por_dia, gastos_por_dia], ignore_index=True)

                import altair as alt
                chart = alt.Chart(comparacion_df).mark_line().encode(
                    x='fecha_dia:T',
                    y='sum(monto):Q',
                    color='tipo:N',
                    tooltip=['fecha_dia:T', 'sum(monto):Q', 'tipo:N']
                ).properties(
                    width='container',
                    title='Comparación de Pagos: Planillas vs Gastos (por día)'
                )

                st.altair_chart(chart, use_container_width=True)

    # Mostrar detalle de pagos recientes
    if not df_pagos.empty:
        df_pagos['razon_social'] = df_pagos['razon_social'].fillna('Desconocido').astype(str).str.strip()
        # Cambiar los nombres de las columnas en la tabla 'Detalle de Pagos Recientes'
        st.markdown("### 📋 Detalle de Pagos Recientes")
        st.dataframe(df_pagos[['fecha', 'tipo_pago', 'campana', 'razon_social', 'monto']].rename(columns={
            'fecha': 'FECHA',
            'tipo_pago': 'TIPO DE PAGO',
            'campana': 'CAMPAÑA',
            'razon_social': 'RAZON SOCIAL',
            'monto': 'MONTO'
        }).assign(
            FECHA=lambda x: x['FECHA'].dt.strftime('%d/%m/%Y'),
            MONTO=lambda x: x['MONTO'].apply(lambda m: f"S/. {m:,.2f}")
        ), use_container_width=True)

# Título principal con icono y tamaño grande
st.markdown(
    """
    <div style='display: flex; align-items: center;'>
        <img src='https://img.icons8.com/ios-filled/64/228B22/safe.png' style='margin-right: 16px;'/>
        <h1 style='display: inline; font-size: 3rem; font-weight: bold;'>Dashboard AFP PRIMA - WORLDTEL 2025</h1>
    </div>
    """,
    unsafe_allow_html=True
)

# Espaciado entre título y subtítulo
st.markdown("<div style='height: 32px;'></div>", unsafe_allow_html=True)



st.markdown("""
<div style='display: flex; align-items: center;'>
    <img src='https://img.icons8.com/color/48/000000/combo-chart.png' style='margin-right: 10px;'/>
    <h3 style='display: inline; font-size: 2rem; margin: 0;'>KPIs</h3>
</div>
""", unsafe_allow_html=True)

# KPIs
total_cuentas = len(df)
monto_deuda = df['DEUDA TOTAL'].sum()
monto_gastos_admin = df['GASTOS ADMIN'].sum()
rec_planillas = df['REC. PLANILLAS'].sum()
rec_gastos = df['REC. GASTOS'].sum()


# % Barrido (clientes gestionados)
casos_barridos = df['ULTIMA FECHA GESTION'].notna().sum()
porcentaje_barrido = (casos_barridos / total_cuentas * 100) if total_cuentas > 0 else 0

# Tarjetas de KPIs

# CSS para ocupar todo el ancho

# CSS para ocupar todo el ancho y animación hover en tarjetas
st.markdown("""
<style>
.kpi-row {
    display: flex;
    flex-wrap: wrap;
    gap: 20px;
    width: 100%;
}
.kpi-card {
    flex: 1 1 0;
    min-width: 220px;
    max-width: 100%;
    background: #fff;
    border-radius: 20px;
    padding: 30px;
    box-shadow: 0 2px 8px rgba(0,0,0,0.07);
    margin: 0;
    transition: transform 0.2s, box-shadow 0.2s;
}
.kpi-card:hover {
    transform: translateY(-8px) scale(1.04);
    box-shadow: 0 8px 24px rgba(0,0,0,0.15);
    z-index: 2;
}
.kpi-card h4 {
    margin: 0 0 10px 0;
}
</style>
""", unsafe_allow_html=True)

st.markdown(f"""
<div class='kpi-row'>
    <div class='kpi-card' style='background: linear-gradient(135deg, #b3d8fd 0%, #6eb6ff 100%);'>
        <h4><img src='https://img.icons8.com/ios-filled/24/000000/bank-cards.png'/> TOTAL CUENTAS</h4>
        <p style='font-size:2.5rem; font-weight:bold; color:#1a4fa3; margin:0;'>{total_cuentas:,}</p>
    </div>
    <div class='kpi-card' style='background: linear-gradient(135deg, #c6f6d5 0%, #68d391 100%);'>
        <h4><span style='font-size:1.2rem;'>💰</span> DEUDA TOTAL</h4>
        <p style='font-size:2rem; font-weight:bold; color:#228b22; margin:0;'>S/. {monto_deuda:,.2f}</p>
    </div>
    <div class='kpi-card' style='background: linear-gradient(135deg, #ffe6b3 0%, #ffb366 100%);'>
        <h4><span style='font-size:1.2rem;'>🏦</span> GASTOS ADMIN</h4>
        <p style='font-size:2rem; font-weight:bold; color:#ff6600; margin:0;'>S/. {monto_gastos_admin:,.2f}</p>
    </div>
</div>
<div style='height: 20px;'></div>
<div class='kpi-row'>
    <div class='kpi-card' style='background: linear-gradient(135deg, #fff9c4 0%, #ffe082 100%);'>
        <h4><img src='https://img.icons8.com/ios-filled/24/000000/combo-chart.png'/> % BARRIDO</h4>
        <p style='font-size:2rem; font-weight:bold; color:#ff9800; margin:0;'>{porcentaje_barrido:.1f}%</p>
    </div>
    <div class='kpi-card' style='background: linear-gradient(135deg, #f8bbd0 0%, #f06292 100%);'>
        <h4><span style='font-size:1.2rem;'>🏦</span> REC. PLANILLAS</h4>
        <p style='font-size:2rem; font-weight:bold; color:#ad1457; margin:0;'>S/. {rec_planillas:,.2f}</p>
    </div>
    <div class='kpi-card' style='background: linear-gradient(135deg, #e1bee7 0%, #ba68c8 100%);'>
        <h4><img src='https://img.icons8.com/ios-filled/24/000000/atm.png'/> REC. GASTOS</h4>
        <p style='font-size:2rem; font-weight:bold; color:#6a1b9a; margin:0;'>S/. {rec_gastos:,.2f}</p>
    </div>
</div>
""", unsafe_allow_html=True)
# ================= TABLA RESUMEN POR CAMPAÑA =================
st.markdown("---")
st.markdown("<h2>📋 Tabla Resumen por Campaña</h2>", unsafe_allow_html=True)


# Agrupar por campaña y calcular los valores, incluyendo gestionados
tabla_campana = df.groupby('CAMPAÑA').agg(
    TOTAL_CUENTAS=('CAMPAÑA', 'count'),
    REC_PLANILLAS=('REC. PLANILLAS', 'sum'),
    REC_GASTOS=('REC. GASTOS', 'sum'),
    DEUDA_TOTAL=('DEUDA TOTAL', 'sum'),
    GASTOS_ADMIN=('GASTOS ADMIN', 'sum'),
    GESTIONADOS=('ULTIMA FECHA GESTION', lambda x: x.notna().sum())
).reset_index()

# % PLANILLAS y % GASTOS ADMIN
tabla_campana['% PLANILLAS'] = np.where(
    tabla_campana['DEUDA_TOTAL'] > 0,
    tabla_campana['REC_PLANILLAS'] / tabla_campana['DEUDA_TOTAL'] * 100,
    0
)
tabla_campana['% GASTOS ADMIN'] = np.where(
    tabla_campana['GASTOS_ADMIN'] > 0,
    tabla_campana['REC_GASTOS'] / tabla_campana['GASTOS_ADMIN'] * 100,
    0
)

# Añadir la columna % BARRIDO
tabla_campana['% BARRIDO'] = np.where(
    tabla_campana['TOTAL_CUENTAS'] > 0,
    tabla_campana['GESTIONADOS'] / tabla_campana['TOTAL_CUENTAS'] * 100,
    0
)

# Formatear montos
tabla_campana['REC_PLANILLAS'] = tabla_campana['REC_PLANILLAS'].apply(lambda x: f"S/. {x:,.2f}")
tabla_campana['REC_GASTOS'] = tabla_campana['REC_GASTOS'].apply(lambda x: f"S/. {x:,.2f}")
tabla_campana['DEUDA_TOTAL'] = tabla_campana['DEUDA_TOTAL'].apply(lambda x: f"S/. {x:,.2f}")
tabla_campana['GASTOS_ADMIN'] = tabla_campana['GASTOS_ADMIN'].apply(lambda x: f"S/. {x:,.2f}")
tabla_campana['% PLANILLAS'] = tabla_campana['% PLANILLAS'].apply(lambda x: f"{float(x.replace('%','')):.2f}%" if isinstance(x, str) else f"{x:.2f}%")
tabla_campana['% GASTOS ADMIN'] = tabla_campana['% GASTOS ADMIN'].apply(lambda x: f"{float(x.replace('%','')):.2f}%" if isinstance(x, str) else f"{x:.2f}%")
tabla_campana['% BARRIDO'] = tabla_campana['% BARRIDO'].apply(lambda x: f"{x:.2f}%")


# Calcular totales para cada columna relevante
totales = {
    'CAMPAÑA': 'TOTAL',
    'TOTAL CUENTAS': tabla_campana['TOTAL_CUENTAS'].sum(),
    'REC PLANILLAS': f"S/. {df['REC. PLANILLAS'].sum():,.2f}",
    'REC GASTOS': f"S/. {df['REC. GASTOS'].sum():,.2f}",
    'DEUDA TOTAL': f"S/. {df['DEUDA TOTAL'].sum():,.2f}",
    'GASTOS ADMIN': f"S/. {df['GASTOS ADMIN'].sum():,.2f}",
    'GESTIONADOS': tabla_campana['GESTIONADOS'].sum(),
    '% PLANILLAS': f"{(df['REC. PLANILLAS'].sum()/df['DEUDA TOTAL'].sum()*100 if df['DEUDA TOTAL'].sum()>0 else 0):.2f}%",
    '% GASTOS ADMIN': f"{(df['REC. GASTOS'].sum()/df['GASTOS ADMIN'].sum()*100 if df['GASTOS ADMIN'].sum()>0 else 0):.2f}%",
    '% BARRIDO': f"{(tabla_campana['GESTIONADOS'].sum()/tabla_campana['TOTAL_CUENTAS'].sum()*100 if tabla_campana['TOTAL_CUENTAS'].sum()>0 else 0):.2f}%"
};

# Renombrar todas las columnas con '_' por ' '
tabla_campana = tabla_campana.rename(columns=lambda x: x.replace('_', ' '))
totales = {k.replace('_', ' '): v for k, v in totales.items()}
column_order = [
    'CAMPAÑA',
    'TOTAL CUENTAS',
    'GESTIONADOS',
    '% BARRIDO',
    'DEUDA TOTAL',
    'REC PLANILLAS',
    'GASTOS ADMIN',
    'REC GASTOS',
    '% PLANILLAS',
    '% GASTOS ADMIN'
]
tabla_campana_totales = pd.concat([
    tabla_campana,
    pd.DataFrame([totales])
], ignore_index=True)

tabla_campana_totales = pd.concat([
         tabla_campana,
         pd.DataFrame([totales])
    ], ignore_index=True)
tabla_campana_totales = tabla_campana_totales[column_order]



# Mejor visual con pandas Styler (tabla no interactiva)
def highlight_totals(row):
    return ['background-color: #e3eafc; color: #1a4fa3; font-weight: bold;' if row.name == len(tabla_campana_totales)-1 else '' for _ in row]

def highlight_percent(val):
    try:
        num = float(str(val).replace('%',''))
        if num > 5:
            return 'background-color: #d4edda; color: #228b22; font-weight: bold;'
        elif num > 0:
            return 'background-color: #fff3cd; color: #ff9800;'
        else:
            return 'background-color: #f8d7da; color: #c82333;'
    except:
        return ''



st.markdown("<hr>", unsafe_allow_html=True)
st.markdown("""
<style>
.tabla-dashboard {
    width: 100%;
    border-collapse: separate;
    border-spacing: 0;
    font-size: 1.05em;
}
.tabla-dashboard th {
    background: #23395d;
    color: #fff;
    font-weight: bold;
    padding: 12px 8px;
    border-radius: 12px 12px 0 0;
    border: none;
}
.tabla-dashboard th:nth-child(1) {
    min-width: 80px; /* Compactar CAMPAÑA */
    max-width: 100px;
}
.tabla-dashboard th:nth-child(2) {
    min-width: 80px; /* Compactar TOTAL CUENTAS */
    max-width: 100px;
}
.tabla-dashboard th:nth-child(3) {
    min-width: 80px; /* Compactar GESTIONADOS */
    max-width: 100px;
}
.tabla-dashboard th:nth-child(4) {
    min-width: 80px; /* Compactar % BARRIDO */
    max-width: 100px;
}
.tabla-dashboard th:nth-child(6) {
    min-width: 150px; /* Reducir el ancho mínimo para REC PLANILLAS */
    max-width: 180px;
}
.tabla-dashboard th:nth-child(7) {
    min-width: 150px; /* Reducir el ancho mínimo para GASTOS ADMIN */
    max-width: 180px;
}
.tabla-dashboard th:nth-child(10) {
    min-width: 120px; /* Reducir el ancho mínimo para % GASTOS ADMIN */
    max-width: 140px;
}
.tabla-dashboard th:nth-child(5) {
    min-width: 200px; /* Reducir el ancho mínimo para DEUDA TOTAL */
    max-width: 220px;
}
.tabla-dashboard td {
    background: #f6f8fa;
    padding: 10px 8px;
    border-bottom: 1px solid #e3eafc;
    text-align: center;
}
.tabla-dashboard tr:last-child td {
    background: #2986cc;
    color: #fff;
    font-weight: bold;
    border-bottom: 2px solid #2986cc;
}
.tabla-dashboard .total {
    background: #2986cc !important;
    color: #fff !important;
    font-weight: bold;
}
.tabla-dashboard .percent-high {
    color: #228b22; font-weight: bold;
}
.tabla-dashboard .percent-low {
    color: #ff9800;
}
.tabla-dashboard .percent-zero {
    color: #c82333;
}
</style>
""", unsafe_allow_html=True)

# Ajustar encabezados para reflejar el nuevo orden de columnas
headers = [
    ("<span style='font-size:1.2em;'>🎯</span> CAMPAÑA"),
    ("<span style='font-size:1.2em;'>💳</span> TOTAL CUENTAS"),
    ("GESTIONADOS"),
    ("<span style='font-size:1.2em;'>🧹</span> % BARRIDO"),
    ("<span style='font-size:1.2em;'>💰</span> DEUDA TOTAL"),
    ("<span style='font-size:1.2em;'>🏦</span> REC PLANILLAS"),
    ("<span style='font-size:1.2em;'>🏦</span> GASTOS ADMIN"),
    ("<span style='font-size:1.2em;'>🏧</span> REC GASTOS"),
    ("<span style='font-size:1.2em;'>📊</span> % PLANILLAS"),
    ("<span style='font-size:1.2em;'>📈</span> % GASTOS ADMIN")
]

# Construir tabla HTML
def percent_class(val):
        try:
                num = float(str(val).replace('%',''))
                if num > 5:
                        return 'percent-high'
                elif num > 0:
                        return 'percent-low'
                else:
                        return 'percent-zero'
        except:
                return ''

tabla_html = "<table class='tabla-dashboard'>"
tabla_html += "<tr>" + "".join([f"<th>{h}</th>" for h in headers]) + "</tr>"

for i, row in tabla_campana_totales.iterrows():
        is_total = (row['CAMPAÑA'] == 'TOTAL')
        tabla_html += "<tr>"
        for col in tabla_campana_totales.columns:
                val = row[col]
                cell_class = "total" if is_total else ""
                if col in ['% PLANILLAS', '% GASTOS ADMIN', '% BARRIDO']:
                        cell_class += " " + percent_class(val)
                tabla_html += f"<td class='{cell_class.strip()}'>{val}</td>"
        tabla_html += "</tr>"
tabla_html += "</table>"

st.markdown(tabla_html, unsafe_allow_html=True)

# ================= GRAFICOS DE PASTEL POR CAMPAÑA =================
st.markdown("---")
st.markdown("""
<div style='display: flex; align-items: center;'>
    <img src='https://img.icons8.com/color/48/000000/pie-chart.png' style='margin-right: 10px;'/>
    <h2 style='display: inline; font-size: 2.2rem; margin: 0;'>Gráficos de Recaudo por Campaña</h2>
</div>
""", unsafe_allow_html=True)

import matplotlib.pyplot as plt


colors = ['#A3CEF1', '#FFB347', '#B5EAD7', '#C7CEEA', '#FFD6E0', '#B28DFF', '#FFB3BA', '#3A86FF']
# Filtrar campañas con monto > 0 para cada gráfico
df_planillas = tabla_campana[tabla_campana['REC PLANILLAS'].apply(lambda x: float(str(x).replace('S/.','').replace(',','')) > 0)]
campanias_planillas = df_planillas['CAMPAÑA'].tolist()
rec_planillas = df_planillas['REC PLANILLAS'].apply(lambda x: float(str(x).replace('S/.','').replace(',',''))).tolist()

df_gastos = tabla_campana[tabla_campana['REC GASTOS'].apply(lambda x: float(str(x).replace('S/.','').replace(',','')) > 0)]
campanias_gastos = df_gastos['CAMPAÑA'].tolist()
rec_gastos = df_gastos['REC GASTOS'].apply(lambda x: float(str(x).replace('S/.','').replace(',',''))).tolist()

palette = ['#A3CEF1', '#FFB347', '#B5EAD7', '#C7CEEA', '#FFD6E0', '#B28DFF', '#FFB3BA', '#3A86FF']
# Asignar colores fijos por nombre de campaña
color_por_campania = {
    'REAL TOTAL': '#FFB347',  # naranja pastel
    'PRESUNTA': '#FFD6E0',   # rosa pastel
    'FLUJO': '#A3CEF1',      # azul pastel
    'REDIRECCIONAMIENTO': '#B5EAD7', # verde pastel
}
# Si hay campañas adicionales, asignarles colores de la paleta en orden
campanias_todas = list(set(campanias_planillas + campanias_gastos))
extra_camps = [c for c in campanias_todas if c not in color_por_campania]
for i, camp in enumerate(extra_camps):
    color_por_campania[camp] = palette[i % len(palette)]

colors_planillas = [color_por_campania.get(camp, '#C7CEEA') for camp in campanias_planillas]
colors_gastos = [color_por_campania.get(camp, '#C7CEEA') for camp in campanias_gastos]

# Mostrar los gráficos uno al lado del otro y más pequeños
st.markdown("""
<style>
.pie-row {
    display: flex;
    flex-direction: row;
    justify-content: center;
    align-items: center;
    gap: 40px;
}
.pie-col {
    flex: 1;
    display: flex;
    flex-direction: column;
    align-items: center;
}
</style>
""", unsafe_allow_html=True)

figsize = (5.5, 5.5)
fig1, ax1 = plt.subplots(figsize=figsize)

# Etiquetas con monto para REC. PLANILLAS
labels_planillas = [f"{campanias_planillas[i]}\nS/. {rec_planillas[i]:,.2f}" for i in range(len(campanias_planillas))]
# Usar los colores definidos por campaña
colors_planillas = [color_por_campania.get(camp, '#C7CEEA') for camp in campanias_planillas]
wedges1, texts1, autotexts1 = ax1.pie(rec_planillas, labels=labels_planillas, autopct='%1.1f%%', colors=colors_planillas, startangle=140)
ax1.set_title('Recaudo de Planillas por Campaña', fontsize=15)
ax1.axis('equal')
plt.tight_layout()

fig2, ax2 = plt.subplots(figsize=figsize)

# Etiquetas con monto para REC. GASTOS
labels_gastos = [f"{campanias_gastos[i]}\nS/. {rec_gastos[i]:,.2f}" for i in range(len(campanias_gastos))]
colors_gastos = [color_por_campania.get(camp, '#C7CEEA') for camp in campanias_gastos]
wedges2, texts2, autotexts2 = ax2.pie(rec_gastos, labels=labels_gastos, autopct='%1.1f%%', colors=colors_gastos, startangle=140)
ax2.set_title('Recaudo de Gastos por Campaña', fontsize=15)
ax2.axis('equal')
plt.tight_layout()

col1, col2 = st.columns(2)
with col1:
    st.pyplot(fig1, use_container_width=True)
with col2:
    st.pyplot(fig2, use_container_width=True)
# ================= FIN GRAFICOS DE PASTEL POR CAMPAÑA =================
# ================= GRAFICOS DE BARRAS HORIZONTALES POR ASESOR =================
# Este bloque muestra dos gráficos de barras horizontales, uno para REC. PLANILLAS y otro para REC. GASTOS por asesor.
st.markdown("---")
st.markdown("""
<div style='display: flex; align-items: center;'>
    <img src='https://img.icons8.com/color/48/000000/bar-chart.png' style='margin-right: 10px;'/>
    <h2 style='display: inline; font-size: 2.2rem; margin: 0;'>Gráficos de Recaudo por Asesor</h2>
</div>
""", unsafe_allow_html=True)

# Agrupar datos por asesor

# Extraer solo el primer nombre del asesor
df['ASESOR_PRIMER_NOMBRE'] = df['ASESOR'].astype(str).apply(lambda x: x.split()[0] if isinstance(x, str) and len(x.split()) > 0 else x)
tabla_asesor = df.groupby('ASESOR_PRIMER_NOMBRE').agg(
    REC_PLANILLAS=('REC. PLANILLAS', 'sum'),
    REC_GASTOS=('REC. GASTOS', 'sum')
).reset_index()

# Ordenar por monto descendente
tabla_asesor_planillas = tabla_asesor.sort_values('REC_PLANILLAS', ascending=True)
tabla_asesor_gastos = tabla_asesor.sort_values('REC_GASTOS', ascending=True)

# Gráfico de barras horizontales para REC. PLANILLAS por asesor
fig_bar_planillas, ax_bar_planillas = plt.subplots(figsize=(7, 5))
ax_bar_planillas.barh(tabla_asesor_planillas['ASESOR_PRIMER_NOMBRE'], tabla_asesor_planillas['REC_PLANILLAS'], color='#FFB347')
ax_bar_planillas.set_xlabel('Recaudo de Planillas (S/.)')
ax_bar_planillas.set_ylabel('Asesor (Primer Nombre)')
ax_bar_planillas.set_title('Recaudo de Planillas por Asesor')
# Etiquetas con monto recaudado en cada barra
for i, (valor, nombre) in enumerate(zip(tabla_asesor_planillas['REC_PLANILLAS'], tabla_asesor_planillas['ASESOR_PRIMER_NOMBRE'])):
    ancho = ax_bar_planillas.get_xlim()[1]
    texto = f"S/. {valor:,.2f}"
    if valor > ancho * 0.15:
        # Si la barra es suficientemente grande, mostrar el texto dentro
        ax_bar_planillas.text(valor/2, i, texto, va='center', ha='center', fontsize=10, color='black', fontweight='bold')
    else:
        # Si la barra es pequeña, mostrar el texto fuera, a la derecha
        ax_bar_planillas.text(valor + ancho*0.01, i, texto, va='center', ha='left', fontsize=10, color='black', fontweight='bold')
plt.tight_layout()

# Gráfico de barras horizontales para REC. GASTOS por asesor
fig_bar_gastos, ax_bar_gastos = plt.subplots(figsize=(7, 5))
ax_bar_gastos.barh(tabla_asesor_gastos['ASESOR_PRIMER_NOMBRE'], tabla_asesor_gastos['REC_GASTOS'], color='#A3CEF1')
ax_bar_gastos.set_xlabel('Recaudo de Gastos (S/.)')
ax_bar_gastos.set_ylabel('Asesor (Primer Nombre)')
ax_bar_gastos.set_title('Recaudo de Gastos por Asesor')
# Etiquetas con monto recaudado en cada barra
for i, (valor, nombre) in enumerate(zip(tabla_asesor_gastos['REC_GASTOS'], tabla_asesor_gastos['ASESOR_PRIMER_NOMBRE'])):
    ancho = ax_bar_gastos.get_xlim()[1]
    texto = f"S/. {valor:,.2f}"
    if valor > ancho * 0.15:
        ax_bar_gastos.text(valor/2, i, texto, va='center', ha='center', fontsize=10, color='black', fontweight='bold')
    else:
        ax_bar_gastos.text(valor + ancho*0.01, i, texto, va='center', ha='left', fontsize=10, color='black', fontweight='bold')
plt.tight_layout()

# Mostrar los gráficos uno al costado del otro
col_bar1, col_bar2 = st.columns(2)
with col_bar1:
    st.pyplot(fig_bar_planillas, use_container_width=True)
with col_bar2:
    st.pyplot(fig_bar_gastos, use_container_width=True)
# ================= FIN GRAFICOS DE BARRAS HORIZONTALES POR ASESOR =================
# ================= TABLA RESUMEN POR ASESOR =================
st.markdown("---")
st.markdown("""
<div style='display: flex; align-items: center;'>
    <img src='https://img.icons8.com/color/48/000000/table.png' style='margin-right: 10px;'/>
    <h2 style='display: inline; font-size: 2.2rem; margin: 0;'>Tabla Resumen por Asesor</h2>
</div>
""", unsafe_allow_html=True)

tabla_resumen_asesor = df.groupby('ASESOR').agg(
    QdeCuentas=('ASESOR', 'count'),
    Gestionados=('ULTIMA FECHA GESTION', lambda x: x.notna().sum()),
    DeudaTotal=('DEUDA TOTAL', 'sum'),
    RecPlanillas=('REC. PLANILLAS', 'sum'),
    GastosAdmin=('GASTOS ADMIN', 'sum'),
    RecGastos=('REC. GASTOS', 'sum')
).reset_index()
tabla_resumen_asesor['%Gestion'] = tabla_resumen_asesor.apply(
    lambda row: f"{int(round(row['Gestionados']/row['QdeCuentas']*100)) if row['QdeCuentas']>0 else 0}%", axis=1)
tabla_resumen_asesor = tabla_resumen_asesor[[
    'ASESOR',
    'QdeCuentas',
    'Gestionados',
    '%Gestion',
    'DeudaTotal',
    'RecPlanillas',
    'GastosAdmin',
    'RecGastos'
]]
tabla_resumen_asesor = tabla_resumen_asesor.sort_values('ASESOR', ascending=False)
for col in ['DeudaTotal', 'RecPlanillas', 'GastosAdmin', 'RecGastos']:
    tabla_resumen_asesor[col] = tabla_resumen_asesor[col].apply(lambda x: f"S/. {int(round(x)):,}" if pd.notnull(x) else "")
st.dataframe(
    tabla_resumen_asesor.style.set_table_styles(
        [{'selector': 'th', 'props': [('text-align', 'center')]}]
    ).set_properties(**{'text-align': 'center'}),
    use_container_width=True,
    hide_index=True
)
# ================= FIN TABLA RESUMEN POR ASESOR =================
# ================= TABLA RESUMEN POR PRIORIDAD =================
# Encabezado con icono
st.markdown("""
<div style='display: flex; align-items: center;'>
    <img src='https://img.icons8.com/color/48/000000/flag.png' style='margin-right: 10px;'/>
    <h2 style='display: inline; font-size: 2.2rem; margin: 0;'>Tabla Resumen por Prioridad</h2>
</div>
""", unsafe_allow_html=True)

# Filtro por campaña


col_filtros1, col_filtros2 = st.columns([2,2])
with col_filtros1:
    campanias = df['CAMPAÑA'].unique().tolist()
    opciones = ['TOTAL'] + campanias
    campania_seleccionada = st.radio('Filtrar por campaña:', opciones, horizontal=True)
with col_filtros2:
    asesores = ['TODOS'] + sorted(df['ASESOR'].dropna().unique().tolist())
    asesor_seleccionado = st.selectbox('Filtrar por asesor:', asesores)

# Aplicar ambos filtros
if campania_seleccionada == 'TOTAL':
    df_filtrado = df.copy()
else:
    df_filtrado = df[df['CAMPAÑA'] == campania_seleccionada]
if asesor_seleccionado != 'TODOS':
    df_filtrado = df_filtrado[df_filtrado['ASESOR'] == asesor_seleccionado]

# Generar tabla resumen por prioridad para la campaña seleccionada
tabla_resumen_prioridad = df_filtrado.groupby('PRIORIDAD').agg(
    QdeCuentas=('PRIORIDAD', 'count'),
    Gestionados=('ULTIMA FECHA GESTION', lambda x: x.notna().sum()),
    DeudaTotal=('DEUDA TOTAL', 'sum'),
    RecPlanillas=('REC. PLANILLAS', 'sum'),
    GastosAdmin=('GASTOS ADMIN', 'sum'),
    RecGastos=('REC. GASTOS', 'sum')
).reset_index()
tabla_resumen_prioridad['%Gestion'] = tabla_resumen_prioridad.apply(
    lambda row: f"{int(round(row['Gestionados']/row['QdeCuentas']*100)) if row['QdeCuentas']>0 else 0}%", axis=1)

# Calcular % REC.PLANILLAS sobre DeudaTotal y % REC.GASTOS sobre GastosAdmin
tabla_resumen_prioridad['%Rec.Planillas'] = tabla_resumen_prioridad.apply(
    lambda row: f"{(row['RecPlanillas']/row['DeudaTotal']*100):.2f}%" if row['DeudaTotal']>0 else "0.00%", axis=1)
tabla_resumen_prioridad['%Rec.Gastos'] = tabla_resumen_prioridad.apply(
    lambda row: f"{(row['RecGastos']/row['GastosAdmin']*100):.2f}%" if row['GastosAdmin']>0 else "0.00%", axis=1)

tabla_resumen_prioridad = tabla_resumen_prioridad[[
    'PRIORIDAD',
    'QdeCuentas',
    'Gestionados',
    '%Gestion',
    'DeudaTotal',
    'RecPlanillas',
    'GastosAdmin',
    'RecGastos',
    '%Rec.Planillas',
    '%Rec.Gastos'
]]
tabla_resumen_prioridad = tabla_resumen_prioridad.sort_values('PRIORIDAD', ascending=False)
for col in ['DeudaTotal', 'RecPlanillas', 'GastosAdmin', 'RecGastos']:
    tabla_resumen_prioridad[col] = tabla_resumen_prioridad[col].apply(lambda x: f"S/. {int(round(x)):,}" if pd.notnull(x) else "")

# Calcular totales
total_qdecuentas = tabla_resumen_prioridad['QdeCuentas'][:-1].astype(str).replace({',':''}, regex=True).astype(int).sum() if 'TOTAL' in tabla_resumen_prioridad['PRIORIDAD'].values else tabla_resumen_prioridad['QdeCuentas'].astype(str).replace({',':''}, regex=True).astype(int).sum()
total_gestionados = tabla_resumen_prioridad['Gestionados'][:-1].astype(str).replace({',':''}, regex=True).astype(int).sum() if 'TOTAL' in tabla_resumen_prioridad['PRIORIDAD'].values else tabla_resumen_prioridad['Gestionados'].astype(str).replace({',':''}, regex=True).astype(int).sum()
total_deuda = tabla_resumen_prioridad['DeudaTotal'][:-1].replace({'S/. ': '', ',': ''}, regex=True).astype(float).sum() if 'TOTAL' in tabla_resumen_prioridad['PRIORIDAD'].values else tabla_resumen_prioridad['DeudaTotal'].replace({'S/. ': '', ',': ''}, regex=True).astype(float).sum()
total_planillas = tabla_resumen_prioridad['RecPlanillas'][:-1].replace({'S/. ': '', ',': ''}, regex=True).astype(float).sum() if 'TOTAL' in tabla_resumen_prioridad['PRIORIDAD'].values else tabla_resumen_prioridad['RecPlanillas'].replace({'S/. ': '', ',': ''}, regex=True).astype(float).sum()
total_gastosadmin = tabla_resumen_prioridad['GastosAdmin'][:-1].replace({'S/. ': '', ',': ''}, regex=True).astype(float).sum() if 'TOTAL' in tabla_resumen_prioridad['PRIORIDAD'].values else tabla_resumen_prioridad['GastosAdmin'].replace({'S/. ': '', ',': ''}, regex=True).astype(float).sum()
total_recgastos = tabla_resumen_prioridad['RecGastos'][:-1].replace({'S/. ': '', ',': ''}, regex=True).astype(float).sum() if 'TOTAL' in tabla_resumen_prioridad['PRIORIDAD'].values else tabla_resumen_prioridad['RecGastos'].replace({'S/. ': '', ',': ''}, regex=True).astype(float).sum()
total_porcentaje = f"{int(round(total_gestionados/total_qdecuentas*100)) if total_qdecuentas>0 else 0}%"
total_recplanillas_deuda = f"{(total_planillas/total_deuda*100):.2f}%" if total_deuda>0 else "0.00%"
total_recgastos_gastosadmin = f"{(total_recgastos/total_gastosadmin*100):.2f}%" if total_gastosadmin>0 else "0.00%"

fila_total = {
    'PRIORIDAD': 'TOTAL',
    'QdeCuentas': f"{int(total_qdecuentas):,}",
    'Gestionados': f"{int(total_gestionados):,}",
    '%Gestion': total_porcentaje,
    'DeudaTotal': f"S/. {int(round(total_deuda)):,}",
    'RecPlanillas': f"S/. {int(round(total_planillas)):,}",
    'GastosAdmin': f"S/. {int(round(total_gastosadmin)):,}",
    'RecGastos': f"S/. {int(round(total_recgastos)):,}",
    '%Rec.Planillas': total_recplanillas_deuda,
    '%Rec.Gastos': total_recgastos_gastosadmin
}
tabla_resumen_prioridad = pd.concat([tabla_resumen_prioridad, pd.DataFrame([fila_total])], ignore_index=True)

# Mostrar tabla estática (no interactiva)

# Estilos para la tabla: encabezado y totales
st.markdown("""
<style>
.tabla-prioridad th {
    background: #23395d !important;
    color: #fff !important;
    font-weight: bold;
    padding: 12px 8px;
    border-radius: 12px 12px 0 0;
    border: none;
}
.tabla-prioridad tr:last-child td {
    background: #ffe082 !important;
    color: #1a4fa3 !important;
    font-weight: bold;
    border-bottom: 2px solid #ffe082;
}
</style>
""", unsafe_allow_html=True)

# Generar HTML con clase personalizada
tabla_html = tabla_resumen_prioridad.to_html(index=False, classes='tabla-prioridad')
st.markdown(tabla_html, unsafe_allow_html=True)
# ================= FIN TABLA RESUMEN POR PRIORIDAD =================

# ================= ANALISIS ESTRATEGICO POR NIVEL DE PRIORIDAD =================
st.markdown("---")
st.markdown("""
<div style='display: flex; align-items: center;'>
    <img src='https://img.icons8.com/color/48/000000/strategy-board.png' style='margin-right: 10px;'/>
    <h2 style='display: inline; font-size: 2.2rem; margin: 0;'>Análisis Estratégico por Nivel de Prioridad</h2>
</div>
""", unsafe_allow_html=True)

# Clasificación de casos
df_analisis = df.copy()
df_analisis['NIVEL_RIESGO'] = 'BAJA'
# CRÍTICO: PRIORIDAD contiene "12" + CONTACTABILIDAD = "Contacto Directo" + TIPO DE PAGO = "FALTA PAGO DE PAGO PLANILLAS"
# CRÍTICO: PRIORIDAD contiene "12" + CONTACTABILIDAD = "Contacto Directo" + TIPO DE PAGO = "FALTA PAGO DE PAGO PLANILLAS"
# CRÍTICO: PRIORIDAD inicia con "12", CONTACTABILIDAD tiene algún valor y REC. PLANILLAS está vacío o cero
cond_critico = (
    df_analisis['PRIORIDAD'].astype(str).str.startswith('13') &
    (df_analisis['CONTACTABILIDAD'].astype(str).str.strip().str.lower() == 'contacto directo') &
    ((df_analisis['REC. PLANILLAS'].isna()) | (df_analisis['REC. PLANILLAS'] == 0) | (df_analisis['REC. PLANILLAS'] == ''))
)
df_analisis['NIVEL_RIESGO'] = 'BAJA'  # Reiniciar para evitar solapamientos
df_analisis.loc[cond_critico, 'NIVEL_RIESGO'] = '+ALTA'
# ALTO: PRIORIDAD contiene "12" pero NO es crítico
cond_alto = (
    df_analisis['PRIORIDAD'].astype(str).str.startswith('13') & (~cond_critico)
)
df_analisis.loc[cond_alto, 'NIVEL_RIESGO'] = 'ALTA'
# MEDIO: PRIORIDAD contiene "11", "10", "09", "08", "07", "06", "05"
cond_medio = df_analisis['PRIORIDAD'].astype(str).str.startswith(('12','11', '10', '09', '08', '07', '06', '05'))
df_analisis.loc[cond_medio & (~cond_critico) & (~cond_alto), 'NIVEL_RIESGO'] = 'MEDIA'
# BAJO: el resto (ya está por defecto)
# MEDIO: PRIORIDAD contiene "11", "10", "09", "08", "07", "06", "05"
cond_medio = df_analisis['PRIORIDAD'].astype(str).str.startswith(('12','11', '10', '09', '08', '07', '06', '05'))
df_analisis.loc[cond_medio, 'NIVEL_RIESGO'] = 'MEDIA'
# BAJO: el resto (ya está por defecto)

# Métricas por nivel
resumen_nivel = df_analisis.groupby('NIVEL_RIESGO').agg(
    CUENTAS=('NIVEL_RIESGO', 'count'),
    DEUDA=('DEUDA TOTAL', 'sum'),
    RECUPERADO=('REC. PLANILLAS', 'sum')
).reset_index()
total_cuentas = resumen_nivel['CUENTAS'].sum()
resumen_nivel['% DEL TOTAL'] = resumen_nivel['CUENTAS'] / total_cuentas * 100

# Ordenar niveles
orden_niveles = ['+ALTA', 'ALTA', 'MEDIA', 'BAJA']
resumen_nivel['ORDEN'] = resumen_nivel['NIVEL_RIESGO'].apply(lambda x: orden_niveles.index(x) if x in orden_niveles else 99)
resumen_nivel = resumen_nivel.sort_values('ORDEN')

# Colores e íconos
iconos = {
    '+ALTA': "<span style='font-size:2.2em;'>🧨</span>",
    'ALTA': "<span style='font-size:2.2em; color:#2e7d32;'>🟢</span>",
    'MEDIA': "<span style='font-size:2.2em; color:#fbc02d;'>🟡</span>",
    'BAJA': "<span style='font-size:2.2em; color:#d32f2f;'>🔴</span>"
}
color_card = {
    '+ALTA': '#66bb6a',
    'ALTA': '#e8f5e9',   # verde claro
    'MEDIA': '#fffde7',  # amarillo claro
    'BAJA': '#ffe6e6'    # rojo claro
}

# Visualización horizontal
st.markdown("""
<style>
.nivel-row {
    display: flex;
    flex-direction: row;
    gap: 32px;
    justify-content: center;
}
.nivel-card {
    flex: 1;
    background: #fff;
    border-radius: 12px;
    padding: 20px;
    box-shadow: 0 2px 8px rgba(0,0,0,0.1);
    text-align: center;
}
.nivel-card h3 {
    margin: 0;
    font-size: 1.5rem;
    font-weight: bold;
}
.nivel-card p {
    margin: 8px 0;
    font-size: 1rem;
}
</style>
""", unsafe_allow_html=True)

cards_html = "<div class='nivel-row'>"
for _, row in resumen_nivel.iterrows():
    nivel = row['NIVEL_RIESGO']
    icono = iconos.get(nivel, '')
    bg = color_card.get(nivel, '#fff')
    cards_html += f"<div class='nivel-card' style='background:{bg};'>"
    cards_html += f"<h3>{icono} {nivel}</h3>"
    cards_html += f"<p><b>Cuentas:</b> {int(row['CUENTAS']):,}</p>"
    cards_html += f"<p><b>Deuda:</b> S/. {int(row['DEUDA']):,}</p>"
    cards_html += f"<p><b>Recuperado:</b> S/. {int(row['RECUPERADO']):,}</p>"
    cards_html += f"<p><b>% del total:</b> {row['% DEL TOTAL']:.1f}%</p>"
    cards_html += "</div>"
cards_html += "</div>"
st.markdown(cards_html, unsafe_allow_html=True)
# ================= FIN ANALISIS ESTRATEGICO POR NIVEL DE PRIORIDAD =================

# Leyenda horizontal y centrada debajo del análisis estratégico
st.markdown("""
<div style='width:100%; display:flex; justify-content:center; margin:24px 0 12px 0;'>
    <div style='display:flex; gap:38px; align-items:center; background:#f7f9fc; border-radius:16px; padding:18px 32px;'>
        <span style='font-size:1.2em;'>⚡ <b>Sistema de Prioridades</b></span>
        <span style='font-size:1.1em;'>🧨 <b>+ALTA:</b> Prioridad 13 + Contacto Directo + Sin Pago</span>
        <span style='font-size:1.1em; color:#388e3c;'>🟢 <b>ALTA:</b> Prioridad 13 (todos)</span>
        <span style='font-size:1.1em; color:#fbc02d;'>🟡 <b>MEDIA:</b> Prioridades 6-12</span>
        <span style='font-size:1.1em; color:#d32f2f;'>🔴 <b>BAJA:</b> Prioridades 1-5</span>
    </div>
</div>
""", unsafe_allow_html=True)

# ================= TABLA DE CASOS CRÍTICO =================

# Función para exportar a Excel
def export_to_excel(df_export):
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from datetime import datetime
    
    output = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Casos Críticos"
    
    # Encabezado con título
    ws['A1'] = "CASOS CRÍTICOS - PRIORIDAD 13 + CONTACTO DIRECTO + SIN PAGO"
    ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="C62828", end_color="C62828", fill_type="solid")
    ws.merge_cells('A1:E1')
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 25
    
    # Fecha de generación
    ws['A2'] = f"Fecha de generación: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
    ws['A2'].font = Font(italic=True, size=10)
    ws.merge_cells('A2:E2')
    ws.row_dimensions[2].height = 18
    
    # Encabezados de columnas
    headers = ['Documento', 'Razón Social', 'Deuda Total', 'Operador', 'Campaña']
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill(start_color="23395D", end_color="23395D", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    
    ws.row_dimensions[4].height = 20
    
    # Datos
    for row_idx, (_, row) in enumerate(df_export.iterrows(), 5):
        ws.cell(row=row_idx, column=1).value = row['Documento']
        ws.cell(row=row_idx, column=2).value = row['Razón Social']
        ws.cell(row=row_idx, column=3).value = row['Deuda Total']
        ws.cell(row=row_idx, column=4).value = row['Operador']
        ws.cell(row=row_idx, column=5).value = row['Campaña']
        
        for col_idx in range(1, 6):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
            if col_idx == 3:  # Alinear números a la derecha
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.alignment = Alignment(horizontal="left")
    
    # Ajustar anchos de columnas
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 18
    
    # Fila de totales
    total_row = len(df_export) + 5
    ws.cell(row=total_row, column=1).value = "TOTAL"
    ws.cell(row=total_row, column=1).font = Font(bold=True)
    ws.cell(row=total_row, column=2).value = len(df_export)
    ws.cell(row=total_row, column=2).font = Font(bold=True)
    for col_idx in range(1, 6):
        cell = ws.cell(row=total_row, column=col_idx)
        cell.fill = PatternFill(start_color="FFE082", end_color="FFE082", fill_type="solid")
        cell.border = thin_border
    
    wb.save(output)
    output.seek(0)
    return output.getvalue()

st.markdown("""
<div style='display: flex; align-items: center; margin-top:32px;'>
    <img src='https://img.icons8.com/emoji/48/000000/bomb-emoji.png' style='margin-right: 10px;'/>
    <h2 style='display: inline; font-size: 2rem; margin: 0; color: #c62828;'>
        Detalle de Casos Críticos (Prioridad 13 + Contacto Directo + Sin Pago)
    </h2>
</div>
""", unsafe_allow_html=True)
df_critico = df_analisis[df_analisis['NIVEL_RIESGO'] == '+ALTA']

# Preparar tabla de casos críticos para mostrar y exportar
cols_critico = {
    'DOCUMENTO': 'Documento',
    'RAZON SOCIAL': 'Razón Social',
    'DEUDA TOTAL': 'Deuda Total',
    'OPERADOR': 'Operador',
    'CAMPAÑA': 'Campaña'
}
df_critico_tabla = df_critico[list(cols_critico.keys())].rename(columns=cols_critico)

st.write(f"Total de casos críticos detectados: {len(df_critico)}")

# Descripción de distribución por campaña
distribucion = df_critico['CAMPAÑA'].value_counts()
desc = "<b>Distribución por Campaña:</b><br>"
for camp, cant in distribucion.items():
    desc += f"• <b>{camp}</b>: {cant} casos<br>"
st.markdown(desc, unsafe_allow_html=True)

# Botón para descargar Excel
if not df_critico.empty:
    excel_data = export_to_excel(df_critico_tabla)
    st.download_button(
        label="📥 Descargar tabla en Excel",
        data=excel_data,
        file_name=f"casos_criticos_{datetime.now().strftime('%d%m%Y_%H%M%S')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="download_criticos"
    )
    st.markdown("<div style='height: 10px;'></div>", unsafe_allow_html=True)

# Estilo para el header de la tabla (rojo más intenso)
st.markdown("""
<style>
.tabla-critico th {
    background: #c62828 !important;
    color: #fff !important;
    font-weight: bold;
    font-size: 1.1em;
    padding: 10px 6px;
    border: none;
}
.tabla-critico td {
    background: #fff;
    color: #222;
    font-size: 1em;
    padding: 8px 6px;
    border-bottom: 1px solid #f3f3f3;
}
</style>
""", unsafe_allow_html=True)

# Leyenda de sistema de prioridades

# Leyenda horizontal y centrada debajo del análisis estratégico

# Mover la leyenda arriba del título de casos críticos


# Mostrar tabla no interactiva, encabezado rojo y scroll horizontal

tabla_html = """
<div style='overflow-x:auto; max-width:100%;'>
    <div style='max-height:340px; overflow-y:auto; border-radius:12px; box-shadow:0 2px 8px rgba(0,0,0,0.07);'>
        <table class='tabla-critico' style='min-width:700px; width:100%;'>
            <thead>
                <tr>
                    <th>Documento</th>
                    <th>Razón Social</th>
                    <th>Deuda Total</th>
                    <th>Operador</th>
                    <th>Campaña</th>
                </tr>
            </thead>
            <tbody>
"""
for _, row in df_critico_tabla.iterrows():
        tabla_html += f"<tr>"
        tabla_html += f"<td>{row['Documento']}</td>"
        tabla_html += f"<td>{row['Razón Social']}</td>"
        tabla_html += f"<td>{row['Deuda Total']}</td>"
        tabla_html += f"<td>{row['Operador']}</td>"
        tabla_html += f"<td>{row['Campaña']}</td>"
        tabla_html += "</tr>"
tabla_html += """
            </tbody>
        </table>
    </div>
</div>
<div style='margin-top:10px; font-weight:bold; color:#c62828;'>Total de casos críticos detectados: {}</div>
""".format(len(df_critico_tabla))
st.markdown(tabla_html, unsafe_allow_html=True)
# === HISTORIAL DE PAGOS (ACTUALIZADO) ===
# Crear df_pagos desde el DataFrame principal, seleccionando columnas que pueden existir
df_pagos = pd.DataFrame()

# Helpers para parseo y limpieza
def _parse_fecha_serie(s):
    # Si vienen como números (serial Excel), convertir desde 1899-12-30
    try:
        if pd.api.types.is_numeric_dtype(s):
            origin = pd.Timestamp('1899-12-30')
            return s.apply(lambda v: origin + pd.to_timedelta(int(v), unit='D') if not pd.isna(v) else pd.NaT)
    except Exception:
        pass
    return pd.to_datetime(s, dayfirst=True, errors='coerce')

def _clean_monto(val):
    try:
        if pd.isna(val):
            return np.nan
        s = str(val)
        # eliminar prefijos tipo 'S/.' y cualquier caracter no numérico salvo ,.-
        s = re.sub(r"S\.?/?\s*", "", s)
        s = re.sub(r"[^0-9,\.-]", "", s)
        if s == '':
            return np.nan
        # si tiene coma y punto, asumimos coma miles y punto decimal -> eliminar comas
        if s.count(',') > 0 and s.count('.') > 0:
            s = s.replace(',', '')
        # si tiene sólo comas, convertir coma decimal a punto
        elif s.count(',') > 0 and s.count('.') == 0:
            s = s.replace(',', '.')
        s = s.replace(' ', '')
        return float(s)
    except Exception:
        return np.nan

# Procesar datos de pagos de planillas y gastos
if not df.empty:
    if 'FECHA DE PAGO P' in df.columns and 'REC. PLANILLAS' in df.columns:
        df_planillas = df[['FECHA DE PAGO P', 'REC. PLANILLAS', 'CAMPAÑA', 'RAZON SOCIAL']].rename(columns={
            'FECHA DE PAGO P': 'fecha',
            'REC. PLANILLAS': 'monto',
            'CAMPAÑA': 'campana',
            'RAZON SOCIAL': 'razon_social'
        }).copy()
        df_planillas['tipo_pago'] = 'PLANILLAS'

    if 'FECHA DE PAGO G' in df.columns and 'REC. GASTOS' in df.columns:
        df_gastos = df[['FECHA DE PAGO G', 'REC. GASTOS', 'CAMPAÑA', 'RAZON SOCIAL']].rename(columns={
            'FECHA DE PAGO G': 'fecha',
            'REC. GASTOS': 'monto',
            'CAMPAÑA': 'campana',
            'RAZON SOCIAL': 'razon_social'
        }).copy()
        df_gastos['tipo_pago'] = 'GASTOS'

    parts = []
    if 'df_planillas' in locals():
        parts.append(df_planillas)
    if 'df_gastos' in locals():
        parts.append(df_gastos)
    if parts:
        df_pagos = pd.concat(parts, ignore_index=True)

# Limpiar y normalizar columnas
if not df_pagos.empty:
    df_pagos['razon_social'] = df_pagos.get('razon_social', '').fillna('Desconocido').astype(str).str.strip()
    df_pagos['campana'] = df_pagos.get('campana', '').fillna('Sin campaña').astype(str).str.strip()
    df_pagos['fecha'] = _parse_fecha_serie(df_pagos['fecha'])
    df_pagos['monto'] = df_pagos['monto'].apply(_clean_monto)
    # Filtrar sólo pagos con monto válido o fecha conocida
    df_pagos = df_pagos.loc[(df_pagos['monto'].notna() & (df_pagos['monto'] > 0)) | df_pagos['fecha'].notna()]

# Verificar si df_pagos contiene datos válidos
if df_pagos.empty:
    st.warning("El DataFrame de pagos está vacío o no contiene pagos recientes con monto/fecha válidos.")
else:
    # Llamar a la función render_historial_pagos con datos limpios
    render_historial_pagos(df_pagos)
# === FIN HISTORIAL DE PAGOS ===